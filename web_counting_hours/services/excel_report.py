# excel_report.py

import tkinter as tk
from tkinter import messagebox, filedialog
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd

def _define_styles():
    """Define e retorna os estilos de célula usados no relatório Excel."""
    styles = {
        'header_fill_gray': PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"),
        'header_fill_light_orange': PatternFill(start_color="FBE2D5", end_color="FBE2D5", fill_type="solid"),
        'header_fill_orange': PatternFill(start_color="F7C7AC", end_color="F7C7AC", fill_type="solid"),
        'light_gray_fill': PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
        'yellow_fill': PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
        'monthly_total_fill': PatternFill(start_color="C5D9F1", end_color="C5D9F1", fill_type="solid"), # Cor azul clara
        'general_total_fill': PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid"), # Cor verde clara
        'no_fill': PatternFill(fill_type=None),
        'bold_font': Font(bold=True),
        'center_align': Alignment(horizontal="center", vertical="center"),
        'right_align': Alignment(horizontal="right", vertical="center"),
        'left_center_align': Alignment(horizontal="left", vertical="center"),
        'thin_border': Border(left=Side(style='thin'), right=Side(style='thin'),
                              top=Side(style='thin'), bottom=Side(style='thin')),
        'no_border': Border()
    }
    return styles

def _add_monthly_total_rows(sheet, current_excel_row, month_start_row, month_num,
                             meses_map, col_ht, col_chd, col_saldo, styles, num_cols_to_apply_style):
    """
    Adiciona 3 linhas de totalização mensal ao sheet do Excel:
    1. Uma linha em branco
    2. A linha "TOTAL MÊS" com as somas
    3. Uma linha em branco (para separar do próximo mês)
    Todas com o preenchimento de fundo de total mensal.
    """
    # Linha 1: Linha em branco antes do total do mês
    for c in range(1, num_cols_to_apply_style + 1):
        cell = sheet.cell(row=current_excel_row, column=c)
        cell.value = ""
        cell.fill = styles['monthly_total_fill']
        cell.border = styles['no_border']
    sheet.row_dimensions[current_excel_row].height = 20
    current_excel_row += 1

    # Linha 2: "TOTAL MÊS" e as somas
    total_month_name = meses_map.get(month_num, "MÊS DESCONHECIDO")
    cell_total_mes = sheet.cell(row=current_excel_row, column=1, value=f"TOTAL {total_month_name}")
    cell_total_mes.font = styles['bold_font']
    cell_total_mes.alignment = styles['left_center_align']
    cell_total_mes.border = styles['no_border']

    for c in range(1, num_cols_to_apply_style + 1):
        cell = sheet.cell(row=current_excel_row, column=c)
        cell.fill = styles['monthly_total_fill']
        if c != 1: # Aplica a borda para as outras colunas se desejar, ou mantém 'no_border'
            cell.border = styles['no_border']

    # Fórmulas de soma
    formula_sum_ht = f"=SUM({get_column_letter(col_ht)}{month_start_row}:{get_column_letter(col_ht)}{current_excel_row - 2})" # Ajuste o fim do range
    cell_ht_sum = sheet.cell(row=current_excel_row, column=col_ht)
    cell_ht_sum.value = formula_sum_ht
    cell_ht_sum.number_format = '[h]:mm;@'
    cell_ht_sum.font = styles['bold_font']
    cell_ht_sum.fill = styles['monthly_total_fill']
    cell_ht_sum.border = styles['no_border']

    formula_sum_chd = f"=SUM({get_column_letter(col_chd)}{month_start_row}:{get_column_letter(col_chd)}{current_excel_row - 2})" # Ajuste o fim do range
    cell_chd_sum = sheet.cell(row=current_excel_row, column=col_chd)
    cell_chd_sum.value = formula_sum_chd
    cell_chd_sum.number_format = '[h]:mm;@'
    cell_chd_sum.font = styles['bold_font']
    cell_chd_sum.fill = styles['monthly_total_fill']
    cell_chd_sum.border = styles['no_border']

    formula_sum_saldo = f"=SUM({get_column_letter(col_saldo)}{month_start_row}:{get_column_letter(col_saldo)}{current_excel_row - 2})" # Ajuste o fim do range
    cell_saldo_sum = sheet.cell(row=current_excel_row, column=col_saldo)
    cell_saldo_sum.value = formula_sum_saldo
    cell_saldo_sum.number_format = '[h]:mm;-h:mm;@'
    cell_saldo_sum.font = styles['bold_font']
    cell_saldo_sum.fill = styles['monthly_total_fill']
    cell_saldo_sum.border = styles['no_border']

    sheet.row_dimensions[current_excel_row].height = 20
    current_excel_row += 1

    # Linha 3: Linha em branco depois do total do mês (para separar do próximo)
    for c in range(1, num_cols_to_apply_style + 1):
        cell = sheet.cell(row=current_excel_row, column=c)
        cell.value = ""
        cell.fill = styles['monthly_total_fill']
        cell.border = styles['no_border']
    sheet.row_dimensions[current_excel_row].height = 20
    current_excel_row += 1

    return current_excel_row

def create_and_fill_excel_report(dataframe: pd.DataFrame, output_filename: str,
                                 employee_name: str, employee_cpf: str, company_cnpj: str,
                                 start_date: str, end_date: str, daily_workload_excel_value: float,
                                 point_columns_names: list, date_column_name: str):
    """
    Cria e preenche um relatório de espelho de ponto em formato Excel.

    Args:
        dataframe (pd.DataFrame): O DataFrame com os dados a serem exportados.
        output_filename (str): O nome do arquivo Excel de saída.
        employee_name (str): Nome do funcionário.
        employee_cpf (str): CPF do funcionário.
        company_cnpj (str): CNPJ da empresa.
        start_date (str): Data inicial do período do relatório.
        end_date (str): Data final do período do relatório.
        daily_workload_excel_value (float): Carga horária diária em formato Excel (fraction of a day).
        point_columns_names (list): Lista de nomes das colunas de ponto no DataFrame.
        date_column_name (str): Nome da coluna original que contém as datas.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = employee_name.split()[0] if employee_name else "Espelho"
    workbook.date_1904 = True
    sheet.sheet_view.showGridLines = False

    styles = _define_styles()

    # --- Criação do Cabeçalho do Relatório ---
    header_data = {
        'A1': "Espelho de Ponto",
        'A2': employee_name,
        'A3': f"CPF: {employee_cpf}",
        'G1': "Lucro+",
        'G2': f"CNPJ: {company_cnpj}",
        'G3': f"{start_date} - {end_date}"
    }

    for cell_ref, value in header_data.items():
        sheet[cell_ref] = value
    print(f"Planilha '{output_filename}' inicializada com o cabeçalho.")

    df_to_export = dataframe.drop(columns=['parsed_date'])

    merge_start_col = 2 # Coluna 'B'
    merge_end_col = 5   # Coluna 'E'
    if point_columns_names:
        try:
            merge_start_col = df_to_export.columns.get_loc(point_columns_names[0]) + 1
            merge_end_col = df_to_export.columns.get_loc(point_columns_names[-1]) + 1
        except KeyError:
            print("Aviso: Algumas colunas de ponto não foram encontradas para definir o intervalo de mesclagem. Usando padrão B-E.")

    start_col_for_new_formulas = len(df_to_export.columns) + 1
    col_idx_horas_trabalhadas = start_col_for_new_formulas
    col_idx_carga_horaria_dia = start_col_for_new_formulas + 1
    col_idx_saldo = start_col_for_new_formulas + 2

    meses_em_portugues = {
        1: "JANEIRO", 2: "FEVEREIRO", 3: "MARÇO", 4: "ABRIL",
        5: "MAIO", 6: "JUNHO", 7: "JULHO", 8: "AGOSTO",
        9: "SETEMBRO", 10: "OUTUBRO", 11: "NOVEMBRO", 12: "DEZEMBRO"
    }

    current_month = None
    month_start_row = 6
    current_excel_row_for_data = 6

    monthly_total_rows_ht = []
    monthly_total_rows_chd = []
    monthly_total_rows_saldo = []

    # --- Inserir Cabeçalho da Tabela ---
    header_row_excel_idx = 5
    header_row_values = list(df_to_export.columns)
    header_row_values.extend(['Horas Trabalhadas', 'C. Hor. Dia', 'Saldo'])

    for j, value in enumerate(header_row_values, start=1):
        cell = sheet.cell(row=header_row_excel_idx, column=j)
        cell.value = value
        cell.font = styles['bold_font']
        cell.alignment = styles['center_align']
        cell.border = styles['thin_border']
        if merge_start_col <= j <= merge_end_col:
            cell.fill = styles['header_fill_gray']
        elif j > len(df_to_export.columns):
            cell.fill = styles['header_fill_orange']
    sheet.row_dimensions[header_row_excel_idx].height = 35

    # Itera sobre as linhas do DataFrame para adicionar os dados
    for df_row_idx, original_df_row in dataframe.iterrows():
        parsed_date = original_df_row['parsed_date']
        if pd.isna(parsed_date):
            continue

        if current_month is None:
            current_month = parsed_date.month
        elif parsed_date.month != current_month:
            # Adiciona as linhas de totalização mensal
            current_excel_row_for_data = _add_monthly_total_rows(
                sheet, current_excel_row_for_data, month_start_row,
                current_month, meses_em_portugues, col_idx_horas_trabalhadas,
                col_idx_carga_horaria_dia, col_idx_saldo, styles, len(header_row_values)
            )
            # Armazena a linha onde a soma mensal foi colocada (para o total geral)
            monthly_total_rows_ht.append(current_excel_row_for_data - 2) # Linha da soma de HT
            monthly_total_rows_chd.append(current_excel_row_for_data - 2) # Linha da soma de CHD
            monthly_total_rows_saldo.append(current_excel_row_for_data - 2) # Linha da soma de Saldo

            month_start_row = current_excel_row_for_data
            current_month = parsed_date.month

        current_row_excel = current_excel_row_for_data
        current_excel_row_for_data += 1

        data_cell_value = str(original_df_row[date_column_name]).lower()
        is_weekend = 'sáb' in data_cell_value or 'dom' in data_cell_value

        point_values_from_df = [str(original_df_row[col_name]) for col_name in point_columns_names]
        valid_points_count = sum(1 for p in point_values_from_df if p != '')
        expected_points = 4

        is_point_row_empty = (valid_points_count == 0)
        has_invalid_point_count = (not is_weekend and valid_points_count != expected_points and valid_points_count > 0)

        for df_col_name in df_to_export.columns:
            excel_col_idx = df_to_export.columns.get_loc(df_col_name) + 1
            cell = sheet.cell(row=current_row_excel, column=excel_col_idx, value=original_df_row[df_col_name])
            cell.alignment = styles['right_align']

            if is_weekend:
                if excel_col_idx >= col_idx_horas_trabalhadas and excel_col_idx <= col_idx_saldo:
                    cell.fill = styles['header_fill_light_orange']
                else:
                    cell.fill = styles['no_fill']
            elif has_invalid_point_count and df_col_name.startswith('Ponto'):
                cell.fill = styles['yellow_fill']
            else:
                cell.fill = styles['no_fill']

        # Preenchimento das 3 colunas de cálculo (Horas Trabalhadas, Carga Horária, Saldo)
        if is_weekend:
            for j in range(col_idx_horas_trabalhadas, col_idx_saldo + 1):
                cell = sheet.cell(row=current_row_excel, column=j)
                cell.value = ""
                cell.fill = styles['header_fill_light_orange']
                cell.alignment = styles['right_align']
        elif is_point_row_empty:
            root = tk.Tk()
            root.withdraw()
            choice = messagebox.askyesnocancel(
                "Linha sem ponto", 
                f"Linha {current_row_excel} ({data_cell_value.upper()}) não possui horários:\n\nSim=Feriado | Não=Folga | Cancelar=Atestado"
            )
            if choice is True: choice = 'F'
            elif choice is False: choice = 'L' 
            else: choice = 'A'
    
            for j in range(col_idx_horas_trabalhadas, col_idx_saldo + 1):
                cell = sheet.cell(row=current_row_excel, column=j)
                cell.value = ""
                cell.fill = styles['header_fill_light_orange']
                cell.alignment = styles['right_align']

            if choice in ['F', 'L', 'A']:
                sheet.merge_cells(start_column=merge_start_col, end_column=merge_end_col,
                                  start_row=current_row_excel, end_row=current_row_excel)
                merged_cell = sheet.cell(row=current_row_excel, column=merge_start_col)
                merged_cell.alignment = styles['center_align']
                merged_cell.fill = styles['light_gray_fill']
                if choice == 'F':
                    merged_cell.value = "FERIADO"
                elif choice == 'L':
                    merged_cell.value = "FOLGA"
                elif choice == 'A':
                    merged_cell.value = "ATESTADO"
            elif choice == 'O':
                try:
                    formula_horas_trabalhadas = (
                        f"=({get_column_letter(df_to_export.columns.get_loc('Ponto 4') + 1)}{current_row_excel} - "
                        f"{get_column_letter(df_to_export.columns.get_loc('Ponto 1') + 1)}{current_row_excel}) - "
                        f"({get_column_letter(df_to_export.columns.get_loc('Ponto 3') + 1)}{current_row_excel} - "
                        f"{get_column_letter(df_to_export.columns.get_loc('Ponto 2') + 1)}{current_row_excel})"
                    )
                except KeyError:
                    print("Aviso: 'Ponto 1', 'Ponto 2', 'Ponto 3' ou 'Ponto 4' não encontrados para a fórmula de Horas Trabalhadas.")
                    formula_horas_trabalhadas = ""

                cell_horas_trabalhadas = sheet.cell(row=current_row_excel, column=col_idx_horas_trabalhadas)
                cell_horas_trabalhadas.value = formula_horas_trabalhadas
                cell_horas_trabalhadas.fill = styles['header_fill_light_orange']
                cell_horas_trabalhadas.number_format = '[h]:mm;@'
                cell_horas_trabalhadas.alignment = styles['right_align']

                cell_carga_horaria_dia = sheet.cell(row=current_row_excel, column=col_idx_carga_horaria_dia)
                cell_carga_horaria_dia.value = daily_workload_excel_value
                cell_carga_horaria_dia.fill = styles['header_fill_light_orange']
                cell_carga_horaria_dia.number_format = '[h]:mm;@'
                cell_carga_horaria_dia.alignment = styles['right_align']

                formula_saldo = (
                    f"={get_column_letter(col_idx_horas_trabalhadas)}{current_row_excel} - "
                    f"{get_column_letter(col_idx_carga_horaria_dia)}{current_row_excel}"
                )
                cell_saldo = sheet.cell(row=current_row_excel, column=col_idx_saldo)
                cell_saldo.value = formula_saldo
                cell_saldo.fill = styles['header_fill_light_orange']
                cell_saldo.number_format = '[h]:mm;-h:mm;@'
                cell_saldo.alignment = styles['right_align']
            else:
                print("Opção inválida. Nenhuma ação tomada para esta linha.")
        elif has_invalid_point_count:
            print(f"Linha {current_row_excel} ({data_cell_value.upper()}) tem {valid_points_count} pontos registrados. Preenchendo em amarelo.")
            try:
                formula_horas_trabalhadas = (
                    f"=({get_column_letter(df_to_export.columns.get_loc('Ponto 4') + 1)}{current_row_excel} - "
                    f"{get_column_letter(df_to_export.columns.get_loc('Ponto 1') + 1)}{current_row_excel}) - "
                    f"({get_column_letter(df_to_export.columns.get_loc('Ponto 3') + 1)}{current_row_excel} - "
                    f"{get_column_letter(df_to_export.columns.get_loc('Ponto 2') + 1)}{current_row_excel})"
                )
            except KeyError:
                print("Aviso: 'Ponto 1', 'Ponto 2', 'Ponto 3' ou 'Ponto 4' não encontrados para a fórmula de Horas Trabalhadas.")
                formula_horas_trabalhadas = ""

            cell_horas_trabalhadas = sheet.cell(row=current_row_excel, column=col_idx_horas_trabalhadas)
            cell_horas_trabalhadas.value = formula_horas_trabalhadas
            cell_horas_trabalhadas.fill = styles['header_fill_light_orange']
            cell_horas_trabalhadas.number_format = '[h]:mm;@'
            cell_horas_trabalhadas.alignment = styles['right_align']

            cell_carga_horaria_dia = sheet.cell(row=current_row_excel, column=col_idx_carga_horaria_dia)
            cell_carga_horaria_dia.value = daily_workload_excel_value
            cell_carga_horaria_dia.fill = styles['header_fill_light_orange']
            cell_carga_horaria_dia.number_format = '[h]:mm;@'
            cell_carga_horaria_dia.alignment = styles['right_align']

            formula_saldo = (
                f"={get_column_letter(col_idx_horas_trabalhadas)}{current_row_excel} - "
                f"{get_column_letter(col_idx_carga_horaria_dia)}{current_row_excel}"
            )
            cell_saldo = sheet.cell(row=current_row_excel, column=col_idx_saldo)
            cell_saldo.value = formula_saldo
            cell_saldo.fill = styles['header_fill_light_orange']
            cell_saldo.number_format = '[h]:mm;-h:mm;@'
            cell_saldo.alignment = styles['right_align']

        else: # Se não for fim de semana, não estiver vazio e tiver EXATAMENTE 4 pontos
            try:
                formula_horas_trabalhadas = (
                    f"=({get_column_letter(df_to_export.columns.get_loc('Ponto 4') + 1)}{current_row_excel} - "
                    f"{get_column_letter(df_to_export.columns.get_loc('Ponto 1') + 1)}{current_row_excel}) - "
                    f"({get_column_letter(df_to_export.columns.get_loc('Ponto 3') + 1)}{current_row_excel} - "
                    f"{get_column_letter(df_to_export.columns.get_loc('Ponto 2') + 1)}{current_row_excel})"
                )
            except KeyError:
                print("Aviso: 'Ponto 1', 'Ponto 2', 'Ponto 3' ou 'Ponto 4' não encontrados para a fórmula de Horas Trabalhadas.")
                formula_horas_trabalhadas = ""

            cell_horas_trabalhadas = sheet.cell(row=current_row_excel, column=col_idx_horas_trabalhadas)
            cell_horas_trabalhadas.value = formula_horas_trabalhadas
            cell_horas_trabalhadas.fill = styles['header_fill_light_orange']
            cell_horas_trabalhadas.number_format = '[h]:mm;@'
            cell_horas_trabalhadas.alignment = styles['right_align']

            cell_carga_horaria_dia = sheet.cell(row=current_row_excel, column=col_idx_carga_horaria_dia)
            cell_carga_horaria_dia.value = daily_workload_excel_value
            cell_carga_horaria_dia.fill = styles['header_fill_light_orange']
            cell_carga_horaria_dia.number_format = '[h]:mm;@'
            cell_carga_horaria_dia.alignment = styles['right_align']

            formula_saldo = (
                f"={get_column_letter(col_idx_horas_trabalhadas)}{current_row_excel} - "
                f"{get_column_letter(col_idx_carga_horaria_dia)}{current_row_excel}"
            )
            cell_saldo = sheet.cell(row=current_row_excel, column=col_idx_saldo)
            cell_saldo.value = formula_saldo
            cell_saldo.fill = styles['header_fill_light_orange']
            cell_saldo.number_format = '[h]:mm;-h:mm;@'
            cell_saldo.alignment = styles['right_align']

        print(f"Processando linha {current_row_excel}...")

    # --- Adicionar Totalização para o ÚLTIMO Mês ---
    if current_month is not None:
        current_excel_row_for_data = _add_monthly_total_rows(
            sheet, current_excel_row_for_data, month_start_row,
            current_month, meses_em_portugues, col_idx_horas_trabalhadas,
            col_idx_carga_horaria_dia, col_idx_saldo, styles, len(header_row_values)
        )
        # Armazena a linha onde a soma mensal foi colocada (para o total geral)
        monthly_total_rows_ht.append(current_excel_row_for_data - 2) # Linha da soma de HT
        monthly_total_rows_chd.append(current_excel_row_for_data - 2) # Linha da soma de CHD
        monthly_total_rows_saldo.append(current_excel_row_for_data - 2) # Linha da soma de Saldo


    # --- Adicionar uma linha em branco (sem preenchimento) para separar a tabela do total geral ---
    current_excel_row_for_data += 1 # Avança para a próxima linha após o último total mensal (já são 3 linhas)
    for c in range(1, len(header_row_values) + 1):
        cell = sheet.cell(row=current_excel_row_for_data, column=c)
        cell.value = ""
        cell.fill = styles['no_fill'] # Linha em branco sem preenchimento
        cell.border = styles['no_border']
    sheet.row_dimensions[current_excel_row_for_data].height = 20
    current_excel_row_for_data += 1


    # --- Adicionar Totalização GERAL ---
    print("\nAdicionando totalização geral...")
    # Linha 1 (vazia para separar) - com preenchimento de total geral
    for c in range(1, len(header_row_values) + 1):
        cell = sheet.cell(row=current_excel_row_for_data, column=c)
        cell.value = ""
        cell.fill = styles['general_total_fill']
        cell.border = styles['no_border']
    sheet.row_dimensions[current_excel_row_for_data].height = 20
    current_excel_row_for_data += 1

    # Linha 2: "TOTAL GERAL" e as somas
    cell_total_geral = sheet.cell(row=current_excel_row_for_data, column=1, value="TOTAL GERAL")
    cell_total_geral.font = styles['bold_font']
    cell_total_geral.alignment = styles['left_center_align']
    cell_total_geral.fill = styles['general_total_fill']
    cell_total_geral.border = styles['no_border']

    # Preenche o fundo para todas as colunas da linha "TOTAL GERAL"
    for c in range(2, len(header_row_values) + 1): # Começa da coluna 2 para não sobrescrever A
        cell = sheet.cell(row=current_excel_row_for_data, column=c)
        cell.fill = styles['general_total_fill']
        cell.border = styles['no_border']

    # Fórmulas de soma do TOTAL GERAL (somando as células de TOTAL MÊS)
    formula_sum_general_ht = f"=SUM({','.join([f'{get_column_letter(col_idx_horas_trabalhadas)}{row}' for row in monthly_total_rows_ht])})"
    cell_general_ht_sum = sheet.cell(row=current_excel_row_for_data, column=col_idx_horas_trabalhadas)
    cell_general_ht_sum.value = formula_sum_general_ht
    cell_general_ht_sum.number_format = '[h]:mm;@'
    cell_general_ht_sum.alignment = styles['right_align']
    cell_general_ht_sum.font = styles['bold_font']
    cell_general_ht_sum.fill = styles['general_total_fill']
    cell_general_ht_sum.border = styles['no_border']

    formula_sum_general_chd = f"=SUM({','.join([f'{get_column_letter(col_idx_carga_horaria_dia)}{row}' for row in monthly_total_rows_chd])})"
    cell_general_chd_sum = sheet.cell(row=current_excel_row_for_data, column=col_idx_carga_horaria_dia)
    cell_general_chd_sum.value = formula_sum_general_chd
    cell_general_chd_sum.number_format = '[h]:mm;@'
    cell_general_chd_sum.alignment = styles['right_align']
    cell_general_chd_sum.font = styles['bold_font']
    cell_general_chd_sum.fill = styles['general_total_fill']
    cell_general_chd_sum.border = styles['no_border']

    formula_sum_general_saldo = f"=SUM({','.join([f'{get_column_letter(col_idx_saldo)}{row}' for row in monthly_total_rows_saldo])})"
    cell_general_saldo_sum = sheet.cell(row=current_excel_row_for_data, column=col_idx_saldo)
    cell_general_saldo_sum.value = formula_sum_general_saldo
    cell_general_saldo_sum.number_format = '[h]:mm;-h:mm;@'
    cell_general_saldo_sum.alignment = styles['right_align']
    cell_general_saldo_sum.font = styles['bold_font']
    cell_general_saldo_sum.fill = styles['general_total_fill']
    cell_general_saldo_sum.border = styles['no_border']

    sheet.row_dimensions[current_excel_row_for_data].height = 20
    current_excel_row_for_data += 1

    # Linha 3 (vazia para separar) - com preenchimento de total geral
    for c in range(1, len(header_row_values) + 1):
        cell = sheet.cell(row=current_excel_row_for_data, column=c)
        cell.value = ""
        cell.fill = styles['general_total_fill']
        cell.border = styles['no_border']
    sheet.row_dimensions[current_excel_row_for_data].height = 20
    current_excel_row_for_data += 1


    # --- Ajuste de Largura das Colunas ---
    print("Ajustando largura das colunas...")
    for col in sheet.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col[:sheet.max_row]:
            try:
                if cell.value is not None:
                    # Se for uma fórmula, não consideramos o comprimento da fórmula em si
                    if not isinstance(cell.value, str) or not cell.value.startswith('='):
                        max_length = max(max_length, len(str(cell.value)))
                    else:
                        max_length = max(max_length, 10) # Tamanho padrão para fórmulas
            except Exception:
                pass
        if max_length > 0:
            sheet.column_dimensions[col_letter].width = max_length + 2
        else:
            sheet.column_dimensions[col_letter].width = 10

    sheet.freeze_panes = 'B6' # Congela painéis

    # --- Finalização e Salvação do Arquivo ---
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        initialname=output_filename
    )
    if file_path:
        workbook.save(file_path)
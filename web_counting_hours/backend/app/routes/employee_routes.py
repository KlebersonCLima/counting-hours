# backend/app/routes/employee_routes.py

from flask import jsonify, request, Blueprint, send_from_directory, send_file
from backend.app.database import get_all_funcionarios, insert_funcionario, delete_funcionario, get_funcionario_por_id, update_funcionario, create_connection
from backend.config import UPLOAD_FOLDER, ALLOWED_EXTENSIONS
from backend.app.utils import allowed_file
import os
from werkzeug.utils import secure_filename
import pandas as pd
from datetime import datetime, timedelta
from io import StringIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import tempfile
import io
import re

employee_bp = Blueprint('employees', __name__)

@employee_bp.route('/', methods=['GET'])
def listar_funcionarios():
    funcionarios = get_all_funcionarios()
    return jsonify(funcionarios)

@employee_bp.route('/', methods=['POST'])
def adicionar_funcionario():
    if 'profile_pic' in request.files:
        file = request.files['profile_pic']
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            filepath = os.path.join(UPLOAD_FOLDER, filename)
            file.save(filepath)
            foto_perfil = filepath  # Salva o caminho do arquivo
        else:
            foto_perfil = None
    else:
        foto_perfil = None

    nome = request.form.get('name')
    cpf = request.form.get('cpf') # Certifique-se de usar 'cpf' se você alterou no HTML
    carga_horaria = request.form.get('carga_horaria')

    if not nome or not cpf or not carga_horaria:
        return jsonify({'error': 'missing_fields'}), 400

    resultado = insert_funcionario(foto_perfil, nome, cpf, carga_horaria)

    if resultado.get('success'):
        return jsonify({'success': True, 'id': resultado['id'], 'message': 'Funcionário cadastrado com sucesso!'}), 201
    else:
        # Tratar diferentes tipos de erro
        error_type = resultado.get('error')
        if error_type == 'duplicate_cpf':
            return jsonify({'error': 'duplicate_cpf', 'message': 'CPF já cadastrado no sistema'}), 409
        elif error_type == 'invalid_hours':
            return jsonify({'error': 'invalid_hours', 'message': 'Carga horária inválida'}), 400
        else:
            return jsonify({'error': 'database_error', 'message': 'Erro interno do servidor'}), 500

@employee_bp.route('/<int:id>', methods=['DELETE'])
def remover_funcionario(id):
    resultado = delete_funcionario(id)
    if resultado.get('success'):
        return jsonify({'success': True, 'message': 'Funcionário removido com sucesso!'}), 200
    else:
        # Tratar diferentes tipos de erro
        error_type = resultado.get('error')
        if error_type == 'not_found':
            return jsonify({'error': 'not_found', 'message': 'Funcionário não encontrado'}), 404
        else:
            return jsonify({'error': 'database_error', 'message': 'Erro interno do servidor'}), 500

@employee_bp.route('/<int:id>', methods=['GET'])
def obter_funcionario(id):
    funcionario = get_funcionario_por_id(id)
    if funcionario:
        return jsonify({'id': funcionario[0], 'foto_perfil': funcionario[1], 'nome': funcionario[2], 'cpf': funcionario[3], 'carga_horaria': funcionario[4]})
    else:
        return jsonify({'message': f'Funcionário com ID {id} não encontrado.'}), 404

@employee_bp.route('/<int:id>', methods=['PUT'])
def atualizar_funcionario(id):
    nome = request.form.get('name')
    cpf = request.form.get('cpf')
    carga_horaria = request.form.get('carga_horaria')

    if not nome or not cpf or not carga_horaria:
        return jsonify({'error': 'missing_fields'}), 400

    # Lógica para lidar com a foto de perfil (se foi atualizada)
    if 'profile_pic' in request.files:
        file = request.files['profile_pic']
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            filepath = os.path.join(UPLOAD_FOLDER, filename)
            file.save(filepath)
            foto_perfil = filepath
        else:
            foto_perfil = request.form.get('old_profile_pic') # Manter a foto antiga se o novo envio for inválido
    else:
        foto_perfil = request.form.get('old_profile_pic') # Manter a foto antiga se não foi enviada uma nova

    resultado = update_funcionario(id, foto_perfil, nome, cpf, carga_horaria)

    if resultado.get('success'):
        return jsonify({'success': True, 'message': 'Funcionário atualizado com sucesso!'}), 200
    else:
        # Tratar diferentes tipos de erro
        error_type = resultado.get('error')
        if error_type == 'duplicate_cpf':
            return jsonify({'error': 'duplicate_cpf', 'message': 'CPF já cadastrado no sistema'}), 409
        elif error_type == 'not_found':
            return jsonify({'error': 'not_found', 'message': 'Funcionário não encontrado'}), 404
        elif error_type == 'invalid_hours':
            return jsonify({'error': 'invalid_hours', 'message': 'Carga horária inválida'}), 400
        else:
            return jsonify({'error': 'database_error', 'message': 'Erro interno do servidor'}), 500

@employee_bp.route('/verificar-cpf', methods=['POST'])
def verificar_cpf():
    """Verifica se um CPF já existe na base de dados"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'status': 'cpf_vazio', 'message': 'CPF não fornecido'}), 400
        
        cpf = data.get('cpf')
        if not cpf:
            return jsonify({'status': 'cpf_vazio', 'message': 'CPF não fornecido'}), 400
        
        # Verificar se o CPF existe
        conn = create_connection()
        if conn is not None:
            try:
                cursor = conn.cursor()
                cursor.execute("SELECT COUNT(*) FROM funcionarios WHERE cpf = ?", (cpf,))
                count = cursor.fetchone()[0]
                exists = count > 0
                
                if exists:
                    return jsonify({
                        'status': 'cpf_existente',
                        'message': 'CPF já existente.',
                        'exists': True
                    }), 200
                else:
                    return jsonify({
                        'status': 'cpf_valido',
                        'message': 'CPF disponível.',
                        'exists': False
                    }), 200
                    
            except Exception as e:
                print(f"Erro ao verificar CPF: {e}")
                return jsonify({
                    'status': 'erro_servidor',
                    'message': 'Erro interno do servidor'
                }), 500
            finally:
                conn.close()
        else:
            return jsonify({
                'status': 'erro_conexao',
                'message': 'Erro de conexão com banco'
            }), 500
            
    except Exception as e:
        print(f"Erro geral na verificação de CPF: {e}")
        return jsonify({
            'status': 'erro_servidor',
            'message': 'Erro interno do servidor'
        }), 500

@employee_bp.route('/profile_pic/<int:id>', methods=['GET'])
def obter_foto_perfil(id):
    funcionario = get_funcionario_por_id(id)
    if funcionario:
        foto_perfil = funcionario[1]
        if foto_perfil and foto_perfil != 'None':
            # Extrair apenas o nome do arquivo do caminho completo
            filename = os.path.basename(foto_perfil)
            return send_from_directory(UPLOAD_FOLDER, filename)
        else:
            return jsonify({'message': 'Foto de perfil não encontrada'}), 404
    else:
        return jsonify({'message': f'Funcionário com ID {id} não encontrado.'}), 404

@employee_bp.route('/processar-csv', methods=['POST'])
def processar_csv_com_justificativas():
    """Processa CSV e identifica dias com pontos incompletos para justificativas"""
    try:
        if 'csv_file' not in request.files:
            return jsonify({'error': 'Arquivo CSV não fornecido'}), 400
        
        csv_file = request.files['csv_file']
        if csv_file.filename == '':
            return jsonify({'error': 'Arquivo CSV não selecionado'}), 400
        
        if not csv_file.filename.lower().endswith('.csv'):
            return jsonify({'error': 'Arquivo deve ser CSV'}), 400
        
        # Salvar arquivo temporariamente
        filename = secure_filename(csv_file.filename)
        temp_path = os.path.join(UPLOAD_FOLDER, f"temp_{filename}")
        csv_file.save(temp_path)
        
        # Obter parâmetros
        start_date = request.form.get('start_date')
        end_date = request.form.get('end_date')
        employee_id = request.form.get('employee_id')
        employee_name = request.form.get('employee_name')
        employee_cpf = request.form.get('employee_cpf')
        employee_hours = request.form.get('employee_hours')
        
        if not start_date or not end_date:
            return jsonify({'error': 'Datas não fornecidas'}), 400
        
        if not employee_id:
            return jsonify({'error': 'Funcionário não selecionado'}), 400
        
        # Converter datas
        start_dt = datetime.strptime(start_date, '%Y-%m-%d')
        end_dt = datetime.strptime(end_date, '%Y-%m-%d')
        
        # Carregar e processar CSV
        try:
            df = pd.read_csv(temp_path)
        except Exception as e:
            return jsonify({'error': f'Erro ao ler CSV: {str(e)}'}), 400
        
        # Identificar colunas de ponto
        point_columns = [col for col in df.columns if col.startswith('Ponto')]
        if not point_columns:
            return jsonify({'error': 'Nenhuma coluna de ponto encontrada'}), 400
        
        # Processar datas do CSV
        date_col = df.columns[0]  # Primeira coluna é a data
        df['parsed_date'] = pd.to_datetime(df[date_col], format='%d/%m/%Y', errors='coerce')
        df = df.dropna(subset=['parsed_date'])
        
        # Identificar dias com pontos incompletos no período
        dias_incompletos = []
        current_date = start_dt
        
        while current_date <= end_dt:
            # Pular finais de semana (sábado=5, domingo=6)
            if current_date.weekday() < 5:  # Segunda a sexta
                # Verificar se o dia existe no CSV
                dia_csv = current_date.strftime('%d/%m/%Y')
                row = df[df[date_col] == dia_csv]
                
                if row.empty:
                    # Dia não encontrado no CSV - considerar como incompleto
                    dias_incompletos.append({
                        'data': current_date.strftime('%d/%m/%Y'),
                        'data_iso': current_date.strftime('%Y-%m-%d'),
                        'motivo': 'dia_ausente'
                    })
                else:
                    # Verificar se há pontos incompletos
                    pontos_vazios = 0
                    for col in point_columns:
                        valor = row[col].iloc[0]
                        if pd.isna(valor) or str(valor).strip() == '' or str(valor).strip().lower() == 'nan':
                            pontos_vazios += 1
                    
                    # Se há pelo menos um ponto vazio, considerar incompleto
                    if pontos_vazios > 0:
                        dias_incompletos.append({
                            'data': current_date.strftime('%d/%m/%Y'),
                            'data_iso': current_date.strftime('%Y-%m-%d'),
                            'motivo': f'ponto_incompleto_{pontos_vazios}_vazios'
                        })
            
            current_date += timedelta(days=1)
        
        # Limpar arquivo temporário
        if os.path.exists(temp_path):
            os.remove(temp_path)
        
        return jsonify({
            'dias_incompletos': dias_incompletos,
            'total_dias_incompletos': len(dias_incompletos),
            'funcionario': {
                'id': employee_id,
                'nome': employee_name,
                'cpf': employee_cpf,
                'carga_horaria': employee_hours
            }
        }), 200
        
    except Exception as e:
        print(f"Erro ao processar CSV: {e}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

@employee_bp.route('/gerar-csv-final', methods=['POST'])
def gerar_csv_final():
    """Gera CSV final com justificativas preenchidas"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Dados não fornecidos'}), 400
        
        csv_content = data.get('csv_content')
        justificativas = data.get('justificativas', {})
        
        if not csv_content:
            return jsonify({'error': 'Conteúdo CSV não fornecido'}), 400
        
        # Processar CSV e adicionar justificativas
        try:
            df = pd.read_csv(StringIO(csv_content))
        except Exception as e:
            print(f"Erro ao ler CSV: {e}")
            return jsonify({'error': f'Erro ao processar CSV: {str(e)}'}), 400
        
        # Adicionar coluna de justificativas se não existir
        if 'Justificativa' not in df.columns:
            df['Justificativa'] = ''
        
        # Preencher justificativas
        for data_iso, justificativa in justificativas.items():
            try:
                # Converter data ISO para formato DD/MM/YYYY
                data_obj = datetime.strptime(data_iso, '%Y-%m-%d')
                data_formatada = data_obj.strftime('%d/%m/%Y')
                
                # Encontrar linha com a data
                date_col = df.columns[0]
                mask = df[date_col] == data_formatada
                
                if mask.any():
                    df.loc[mask, 'Justificativa'] = justificativa
                else:
                    print(f"Data {data_formatada} não encontrada no CSV")
            except Exception as e:
                print(f"Erro ao processar data {data_iso}: {e}")
                continue
        
        # Gerar arquivo CSV
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_filename = f"relatorio_completo_{timestamp}.csv"
        output_path = os.path.join(UPLOAD_FOLDER, output_filename)
        
        try:
            df.to_csv(output_path, index=False, encoding='utf-8')
        except Exception as e:
            print(f"Erro ao salvar CSV: {e}")
            return jsonify({'error': f'Erro ao salvar arquivo: {str(e)}'}), 500
        
        # Retornar arquivo para download
        try:
            return send_from_directory(
                UPLOAD_FOLDER, 
                output_filename, 
                as_attachment=True,
                download_name=output_filename
            )
        except Exception as e:
            print(f"Erro ao enviar arquivo: {e}")
            return jsonify({'error': f'Erro ao enviar arquivo: {str(e)}'}), 500
        
    except Exception as e:
        print(f"Erro geral ao gerar CSV final: {e}")
        return jsonify({'error': f'Erro interno do servidor: {str(e)}'}), 500

@employee_bp.route('/gerar-planilha-final', methods=['POST'])
def gerar_planilha_final():
    """Gera planilha Excel formatada com justificativas preenchidas"""
    try:
        print("DEBUG: Iniciando geracao de planilha final...")
        data = request.get_json()
        if not data:
            print("DEBUG: Dados nao fornecidos")
            return jsonify({'error': 'Dados não fornecidos'}), 400
        
        csv_content = data.get('csv_content')
        justificativas = data.get('justificativas', {})
        employee_name = data.get('employee_name', 'Funcionário')
        employee_cpf = data.get('employee_cpf', '')
        start_date = data.get('start_date', '')
        end_date = data.get('end_date', '')
        
        print(f"DEBUG: Justificativas recebidas: {justificativas}")
        print(f"DEBUG: Funcionario: {employee_name}")
        print(f"DEBUG: Tamanho do CSV: {len(csv_content) if csv_content else 0} caracteres")
        
        if not csv_content:
            print("DEBUG: Conteudo CSV nao fornecido")
            return jsonify({'error': 'Conteúdo CSV não fornecido'}), 400
        
        # Processar CSV e adicionar justificativas
        try:
            print("DEBUG: Processando CSV...")
            
            # Limpar o CSV antes de processar (remover quebras de linha extras e textos)
            cleaned_csv = limpar_csv_content(csv_content)
            print(f"DEBUG: CSV limpo, tamanho: {len(cleaned_csv)} caracteres")
            
            df = pd.read_csv(io.StringIO(cleaned_csv))
            print(f"DEBUG: CSV processado, shape: {df.shape}")
            print(f"DEBUG: Colunas originais: {list(df.columns)}")
            
            # Normalizar nomes das colunas para garantir compatibilidade
            df.columns = [col.strip().title().replace('  ', ' ').replace('Ponto1', 'Ponto 1').replace('Ponto2', 'Ponto 2').replace('Ponto3', 'Ponto 3').replace('Ponto4', 'Ponto 4') for col in df.columns]
            print(f"DEBUG: Colunas normalizadas: {list(df.columns)}")
            print(f"DEBUG: Primeiras linhas do DataFrame:\n{df.head()}\n")
            
            # Verificar se as colunas estão corretas
            if len(df.columns) < 5:
                print("DEBUG: CSV pode ter formato incorreto, tentando ajustar...")
                # Tentar ler com separador diferente
                df = pd.read_csv(io.StringIO(cleaned_csv), sep=',', engine='python')
                df.columns = [col.strip().title().replace('  ', ' ').replace('Ponto1', 'Ponto 1').replace('Ponto2', 'Ponto 2').replace('Ponto3', 'Ponto 3').replace('Ponto4', 'Ponto 4') for col in df.columns]
                print(f"DEBUG: CSV reprocessado, shape: {df.shape}")
                print(f"DEBUG: Colunas reprocessadas: {list(df.columns)}")
                print(f"DEBUG: Primeiras linhas do DataFrame reprocessado:\n{df.head()}\n")
                
        except Exception as e:
            print(f"DEBUG: Erro ao ler CSV: {e}")
            return jsonify({'error': f'Erro ao processar CSV: {str(e)}'}), 400
        
        # Adicionar coluna de justificativas se não existir
        if 'Justificativa' not in df.columns:
            df['Justificativa'] = ''
        
        # Preencher justificativas
        print("DEBUG: Preenchendo justificativas...")
        for data_iso, justificativa in justificativas.items():
            try:
                # Converter data ISO para formato DD/MM/YYYY
                data_obj = datetime.strptime(data_iso, '%Y-%m-%d')
                data_formatada = data_obj.strftime('%d/%m/%Y')
                
                # Encontrar linha com a data
                date_col = df.columns[0]
                mask = df[date_col] == data_formatada
                
                if mask.any():
                    df.loc[mask, 'Justificativa'] = justificativa
                    print(f"DEBUG: Justificativa '{justificativa}' adicionada para {data_formatada}")
                else:
                    print(f"DEBUG: Data {data_formatada} nao encontrada no CSV")
            except Exception as e:
                print(f"DEBUG: Erro ao processar data {data_iso}: {e}")
                continue
        
        # Gerar arquivo Excel formatado
        print("DEBUG: Gerando arquivo Excel...")
        excel_file = criar_planilha_excel_formatada(df, employee_name, employee_cpf, start_date, end_date)
        print(f"DEBUG: Arquivo Excel criado: {excel_file}")
        
        # Verificar se o arquivo existe e tem conteúdo
        if os.path.exists(excel_file):
            file_size = os.path.getsize(excel_file)
            print(f"DEBUG: Arquivo existe, tamanho: {file_size} bytes")
        else:
            print("DEBUG: Arquivo nao existe!")
            return jsonify({'error': 'Arquivo Excel não foi criado'}), 500
        
        # Retornar arquivo para download
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"relatorio_completo_{timestamp}.xlsx"
        print(f"DEBUG: Nome do arquivo para download: {filename}")
        
        return send_file(
            excel_file,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"DEBUG: Erro geral ao gerar planilha final: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Erro interno do servidor: {str(e)}'}), 500

def limpar_csv_content(csv_content):
    """Limpa o conteúdo do CSV removendo quebras de linha extras e textos desnecessários"""
    lines = csv_content.split('\n')
    cleaned_lines = []
    
    for line in lines:
        if line.strip():  # Remove linhas vazias
            # Remove quebras de linha dentro dos campos e textos como "Entrada"/"Saída"
            cleaned_line = line.replace('\n', ' ').replace('\r', ' ')
            
            # Remove textos como "Entrada" e "Saída" dos campos de ponto
            cleaned_line = re.sub(r'(\d{1,2}:\d{2})\s*(?:Entrada|Saída)', r'\1', cleaned_line)
            
            cleaned_lines.append(cleaned_line)
    
    return '\n'.join(cleaned_lines)

def criar_planilha_excel_formatada(df, employee_name, employee_cpf, start_date, end_date):
    """Cria uma planilha Excel corporativa formatada com dados fictícios de 30 dias"""
    
    print(f"DEBUG: Criando planilha com dados ficticios para 30 dias")
    
    # Criar workbook e worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Controle de Ponto"
    
    # Definir estilos corporativos
    styles = {
        'header_fill': PatternFill(start_color="366092", end_color="366092", fill_type="solid"),
        'header_font': Font(bold=True, color="FFFFFF", size=12),
        'subheader_fill': PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
        'subheader_font': Font(bold=True, size=11),
        'data_font': Font(size=10),
        'bold_font': Font(bold=True, size=10),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        ),
        'center_align': Alignment(horizontal="center", vertical="center"),
        'left_align': Alignment(horizontal="left", vertical="center"),
        'right_align': Alignment(horizontal="right", vertical="center"),
        'justificativa_fill': PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid"),
        'feriado_fill': PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid"),
        'folga_fill': PatternFill(start_color="98FB98", end_color="98FB98", fill_type="solid"),
        'atestado_fill': PatternFill(start_color="FFA07A", end_color="FFA07A", fill_type="solid"),
        'monthly_total_fill': PatternFill(start_color="C5D9F1", end_color="C5D9F1", fill_type="solid"),
        'general_total_fill': PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")
    }
    
    # Cabeçalho principal
    ws.merge_cells('A1:I1')
    header_cell = ws['A1']
    header_cell.value = "CONTROLE DE PONTO"
    header_cell.font = Font(bold=True, size=18, color="FFFFFF")
    header_cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    header_cell.alignment = styles['center_align']
    ws.row_dimensions[1].height = 35
    
    # Informações da empresa
    ws['A3'] = "EMPRESA:"
    ws['B3'] = "LUCRO MAIS"
    ws['A4'] = "CNPJ:"
    ws['B4'] = "24.158.957/0000-46"
    ws['A5'] = "PERÍODO:"
    ws['B5'] = f"{start_date} a {end_date}"
    
    # Informações do funcionário
    ws['A7'] = "FUNCIONÁRIO:"
    ws['B7'] = employee_name.upper()
    ws['A8'] = "CPF:"
    ws['B8'] = employee_cpf
    
    # Aplicar estilos às informações
    for row in [3, 4, 5, 7, 8]:
        ws[f'A{row}'].font = Font(bold=True, size=11)
        ws[f'B{row}'].font = Font(size=11)
        ws[f'A{row}'].fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    
    # Cabeçalho da tabela (linha 10)
    header_row = 10
    col_names = ['DATA', 'PONTO 1', 'PONTO 2', 'PONTO 3', 'PONTO 4', 'HORAS TRABALHADAS', 'CARGA HORÁRIA DO DIA', 'SALDO', 'JUSTIFICATIVA']
    
    for col_idx, col_name in enumerate(col_names, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.border = styles['border']
        cell.alignment = styles['center_align']
    
    ws.row_dimensions[header_row].height = 25
    
    # Gerar dados fictícios para 30 dias
    data_start_row = header_row + 1
    
    # Definir datas de justificativas (6 dias: 3 atestados e 3 feriados)
    from datetime import datetime, timedelta
    
    # Converter start_date para datetime
    start_dt = datetime.strptime(start_date, '%Y-%m-%d')
    
    # Dias com justificativas (exemplo: dias 5, 12, 19 para atestados e 8, 15, 22 para feriados)
    atestado_days = [5, 12, 19]  # Dias do mês com atestado
    feriado_days = [8, 15, 22]   # Dias do mês com feriado
    
    # Carga horária diária (8:48 = 8.8 horas)
    daily_workload = 8.8
    
    print(f"DEBUG: Gerando 30 dias de dados ficticios...")
    
    for day in range(30):
        current_date = start_dt + timedelta(days=day)
        row_idx = data_start_row + day
        
        # Verificar se é dia de justificativa
        day_of_month = current_date.day
        justificativa = ""
        row_fill = None
        
        if day_of_month in atestado_days:
            justificativa = "ATESTADO"
            row_fill = styles['atestado_fill']
        elif day_of_month in feriado_days:
            justificativa = "FERIADO"
            row_fill = styles['feriado_fill']
        
        # Determinar se é dia útil (segunda a sexta, exceto feriados)
        is_weekend = current_date.weekday() >= 5  # 5 = sábado, 6 = domingo
        is_feriado = day_of_month in feriado_days
        is_atestado = day_of_month in atestado_days
        
        # Preencher dados da linha
        for col_idx, col_name in enumerate(col_names, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = styles['border']
            cell.font = styles['data_font']
            
            # Aplicar cor de fundo se houver justificativa
            if row_fill:
                cell.fill = row_fill
            
            if col_idx == 1:  # Data
                cell.value = current_date.strftime('%d/%m/%Y')
                cell.alignment = styles['center_align']
                cell.font = styles['bold_font']
            elif col_idx == 2:  # Ponto 1
                if is_weekend or is_feriado or is_atestado:
                    cell.value = ""
                else:
                    cell.value = "08:00"
                cell.alignment = styles['center_align']
            elif col_idx == 3:  # Ponto 2
                if is_weekend or is_feriado or is_atestado:
                    cell.value = ""
                else:
                    cell.value = "12:00"
                cell.alignment = styles['center_align']
            elif col_idx == 4:  # Ponto 3
                if is_weekend or is_feriado or is_atestado:
                    cell.value = ""
                else:
                    cell.value = "13:00"
                cell.alignment = styles['center_align']
            elif col_idx == 5:  # Ponto 4
                if is_weekend or is_feriado or is_atestado:
                    cell.value = ""
                else:
                    cell.value = "17:48"
                cell.alignment = styles['center_align']
            elif col_idx == 6:  # Horas Trabalhadas
                if is_weekend or is_feriado or is_atestado:
                    cell.value = ""
                else:
                    # Fórmula: (Ponto 4 - Ponto 1) - (Ponto 3 - Ponto 2)
                    ponto1_col = get_column_letter(2)  # Ponto 1
                    ponto2_col = get_column_letter(3)  # Ponto 2
                    ponto3_col = get_column_letter(4)  # Ponto 3
                    ponto4_col = get_column_letter(5)  # Ponto 4
                    
                    formula = f"=({ponto4_col}{row_idx}-{ponto1_col}{row_idx})-({ponto3_col}{row_idx}-{ponto2_col}{row_idx})"
                    cell.value = formula
                cell.number_format = '[h]:mm;@'
                cell.alignment = styles['right_align']
            elif col_idx == 7:  # Carga Horária do Dia
                if is_weekend or is_feriado or is_atestado:
                    cell.value = ""
                else:
                    cell.value = daily_workload
                cell.number_format = '[h]:mm;@'
                cell.alignment = styles['right_align']
            elif col_idx == 8:  # Saldo
                if is_weekend or is_feriado or is_atestado:
                    cell.value = ""
                else:
                    # Fórmula: Horas Trabalhadas - Carga Horária
                    ht_col = get_column_letter(6)  # Horas Trabalhadas
                    ch_col = get_column_letter(7)  # Carga Horária
                    
                    formula = f"={ht_col}{row_idx}-{ch_col}{row_idx}"
                    cell.value = formula
                cell.number_format = '[h]:mm;@'
                cell.alignment = styles['right_align']
            elif col_idx == 9:  # Justificativa
                cell.value = justificativa
                cell.alignment = styles['center_align']
    
    # Adicionar totais mensais
    total_row = data_start_row + 30
    ws.cell(row=total_row, column=1, value="TOTAL MENSAL").font = styles['bold_font']
    ws.cell(row=total_row, column=1).fill = styles['monthly_total_fill']
    ws.cell(row=total_row, column=1).border = styles['border']
    ws.cell(row=total_row, column=1).alignment = styles['center_align']
    
    # Fórmula para total de horas trabalhadas
    ht_start = get_column_letter(6) + str(data_start_row)
    ht_end = get_column_letter(6) + str(data_start_row + 29)
    ws.cell(row=total_row, column=6, value=f"=SUM({ht_start}:{ht_end})").number_format = '[h]:mm;@'
    ws.cell(row=total_row, column=6).font = styles['bold_font']
    ws.cell(row=total_row, column=6).fill = styles['monthly_total_fill']
    ws.cell(row=total_row, column=6).border = styles['border']
    ws.cell(row=total_row, column=6).alignment = styles['right_align']
    
    # Fórmula para total de carga horária
    ch_start = get_column_letter(7) + str(data_start_row)
    ch_end = get_column_letter(7) + str(data_start_row + 29)
    ws.cell(row=total_row, column=7, value=f"=SUM({ch_start}:{ch_end})").number_format = '[h]:mm;@'
    ws.cell(row=total_row, column=7).font = styles['bold_font']
    ws.cell(row=total_row, column=7).fill = styles['monthly_total_fill']
    ws.cell(row=total_row, column=7).border = styles['border']
    ws.cell(row=total_row, column=7).alignment = styles['right_align']
    
    # Fórmula para saldo total
    saldo_start = get_column_letter(8) + str(data_start_row)
    saldo_end = get_column_letter(8) + str(data_start_row + 29)
    ws.cell(row=total_row, column=8, value=f"=SUM({saldo_start}:{saldo_end})").number_format = '[h]:mm;@'
    ws.cell(row=total_row, column=8).font = styles['bold_font']
    ws.cell(row=total_row, column=8).fill = styles['monthly_total_fill']
    ws.cell(row=total_row, column=8).border = styles['border']
    ws.cell(row=total_row, column=8).alignment = styles['right_align']
    
    # Aplicar bordas e estilos ao total
    for col in range(2, 10):
        if col != 6 and col != 7 and col != 8:  # Pular colunas com fórmulas
            ws.cell(row=total_row, column=col).fill = styles['monthly_total_fill']
            ws.cell(row=total_row, column=col).border = styles['border']
    
    # Ajustar largura das colunas
    column_widths = [12, 12, 12, 12, 12, 18, 20, 12, 15]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Salvar arquivo temporário
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()
    
    print(f"DEBUG: Planilha salva em: {temp_file.name}")
    return temp_file.name